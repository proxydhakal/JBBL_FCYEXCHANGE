from django.utils import timezone
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
import requests
from django.views import View
import pandas as pd
from django.http import HttpResponse
from .models import FCYRateMaster, FCYExchangeRate, FCYExchangeRequestMaster, FCYDenoMasterTable, CurrencyTable
from apps.branches.models import Branches
from apps.accounts.models import UserAccount
import datetime
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import View
from django.contrib import messages
from apps.core.forms import FCYExchangeRateUploadForm, FCYDenoMasterTableForm, FCYExchangeRequestMasterForm, DenoApprovedForm
from decouple import config
from django.forms import formset_factory
from requests.packages import urllib3
from django.core.mail import send_mail
from django.conf import settings
from num2words import num2words
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.db.models import F, Subquery, OuterRef, ExpressionWrapper, CharField, Value
from django.db.models.functions import Concat
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404
from collections import defaultdict
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Table, TableStyle
from reportlab.pdfgen import canvas
from django.http import FileResponse
from django.contrib.staticfiles.finders import find
from openpyxl import Workbook
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.conf import settings
from django.urls import reverse
from urllib.parse import urljoin
from django.conf import settings


urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
system_date = datetime.date.today()
system_time = datetime.datetime.now().time()
current_datetime = datetime.datetime.now()


def home(request):
    template_name = 'core/index.html'
    if request.user.is_authenticated:
        return redirect('/dashboard/')
    else:
        return render(request, template_name)


@login_required
def dashboard(request):
    template_name = 'core/dashboard/dashboard.html'
    list_times = FCYRateMaster.objects.filter(date=system_date)
    filtered_data = None
    api_data = None 

    try:
        last_object = FCYRateMaster.objects.latest('id')
        filtered_data = FCYExchangeRate.objects.filter(masterid=last_object.pk)
        api_url = f"https://www.nrb.org.np/api/forex/v1/rates?page=1&per_page=20&from={system_date}&to={system_date}"
        response = requests.get(api_url)

        if response.status_code == 200:
            api_data = response.json()
    except ObjectDoesNotExist:
        last_object = None
        alert_message = "No data available in FCYRateMaster"
        context = {
            'filtered_data': filtered_data,
            'list_times': list_times,
            'last_object': last_object,
            'api_data': api_data,  # Ensure api_data is included in the context here
            'alert_message': alert_message
        }
        return render(request, template_name, context)

    context = {
        'filtered_data': filtered_data,
        'list_times': list_times,
        'api_data': api_data,
        'last_object': last_object
    }
    return render(request, template_name, context)


class FCYExchangeRequestCreateView(LoginRequiredMixin, View):
    # Update with your template path
    template_name = 'core/dashboard/create_fcy_exchnage.html'

    def get(self, request):
        form = FCYExchangeRequestMasterForm()
        DenoFormSet = formset_factory(FCYDenoMasterTableForm, extra=1)
        formset = DenoFormSet(request.POST or None)
        branches = Branches.objects.filter(Status='T', EmailStatus='T')
        currencies = CurrencyTable.objects.all()
        return render(request, self.template_name, {'form': form, 'formset': formset, 'branches': branches, 'currencies': currencies})

    def post(self, request):
        form = FCYExchangeRequestMasterForm(request.POST)
        DenoFormSet = formset_factory(FCYDenoMasterTableForm, extra=1)
        formset = DenoFormSet(request.POST or None)
        branches = Branches.objects.filter(Status='T', EmailStatus='T')
        currencies = CurrencyTable.objects.all()

        if form.is_valid() and formset.is_valid():
            preferredBranch = form.cleaned_data['preferredBranch']
            totalEquivalentNPR = form.cleaned_data['totalEquivalentNPR']
            totalEquivalentNPRToWords = form.cleaned_data['totalEquivalentNPRToWords']

            request_master = FCYExchangeRequestMaster.objects.create(
                date=system_date,
                preferredBranch=preferredBranch,
                totalEquivalentNPR=totalEquivalentNPR,
                totalEquivalentNPRToWords=totalEquivalentNPRToWords,
                status='Requested',
                enteredBy=request.user.email,
                enterDate=current_datetime,
                updatedBy='-',
                updateDate=current_datetime,
                deletedBy='-',
                deletedDate=current_datetime
            )

            if request_master:
                # Process and save the formset data
                for deno_form in formset:
                    if deno_form.is_valid():
                        FCYDenoMasterTable.objects.create(
                            masterid=request_master.pk,
                            currency=deno_form.cleaned_data['currency'],
                            unit=deno_form.cleaned_data['unit'],
                            deno=deno_form.cleaned_data['deno'],
                            rate=deno_form.cleaned_data['rate'],
                            equivalentNPR=deno_form.cleaned_data['equivalentNPR'],
                            status='Requested',
                            enteredBy=request.user.email,
                            enterDate=current_datetime,
                            updatedBy='-',
                            updateDate=current_datetime,
                            deletedBy='-',
                            deletedDate=current_datetime,
                        )
                branchdetail = Branches.objects.filter(BranchCode=request_master.preferredBranch).get() 
                context = {
                    'name': f'{request.user.first_name} {request.user.last_name}',
                    'refrenceid': request_master.refrenceid,
                    'branch': branchdetail.BranchName,
                    'datetime': request_master.enterDate,
                    'totalnpr':request_master.totalEquivalentNPR,
                    'email':request.user.email,
                    'company':request.user.company,
                    'branchemail':branchdetail.EmailAddress,
                    'link':urljoin(request.build_absolute_uri('/'), reverse('fcyexchangerequest'))
                }
                html_content = render_to_string('core/email/fcy_request_confirmation.html', context)
                branch_content = render_to_string('core/email/fcy_request_confirmation_branch.html', context)
                
                email = EmailMultiAlternatives('FCY Exchange Confirmation', 'FCY Exchange Confirmation',
                                                settings.EMAIL_HOST_USER, [request.user.email,'shekhar.dhakal@jbbl.com.np', ])
                email.attach_alternative(html_content, 'text/html')
                
                branchemail = EmailMultiAlternatives('FCY Exchange Notification', 'FCY Exchange Notification',
                                                settings.EMAIL_HOST_USER, [request.user.email,'shekhar.dhakal@jbbl.com.np', ])
                branchemail.attach_alternative(branch_content, 'text/html')
                
                email.send(fail_silently=True)
                branchemail.send(fail_silently=True)
                
                messages.success(
                    request, "Your FCY Exchange has been successfully created. Please visit the requested branch!")
                return redirect('/fcyexchangerequest/')

        return render(request, self.template_name, {'form': form, 'formset': formset, 'branches': branches, 'currencies': currencies})


class BranchListView(LoginRequiredMixin, View):
    template_name = 'core/dashboard/branches.html'  

    def get(self, request):
        api_url = 'https://jbbl.com.np/rest-api/forexapi/getbranch/'

        api_data = {
            'username': config('API_USERNAME'),
            'password': config('API_PASSWORD')
        }

        try:
            response = requests.post(api_url, data=api_data, verify=False)
            response.raise_for_status()  

            if response.status_code == 200:
                data = response.json()
                branches = data.get('Branches', [])
                context = {'branches': branches}
                return render(request, self.template_name, context)
            else:
                error_message = f"Failed to fetch branch data from the API. Status code: {response.status_code}"
                return render(request, self.template_name, {'error_message': error_message})
        except requests.exceptions.RequestException as e:
            error_message = f"Request failed: {e}"
            return render(request, self.template_name, {'error_message': error_message})


class FCYExchangeRateUploadView(LoginRequiredMixin, View):
    template_name = 'core/dashboard/uploadfcyrate.html'

    def get(self, request):
        form = FCYExchangeRateUploadForm()
        if(request.user.is_superuser == True):
            return render(request, self.template_name, {'form': form})
        else:
            messages.error(self.request, "Unauthorized User")
            return redirect('/dashboard/')

    def post(self, request):
        form = FCYExchangeRateUploadForm(request.POST, request.FILES)
        if(request.user.is_superuser == True):
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                df = pd.read_excel(excel_file)
                column_names = df.columns.tolist()

                print("Column Names:")
                print(column_names)

                if (len(column_names) == 7 and column_names[0] == 'Currency Code' and column_names[1] == 'Currency' and
                column_names[2] == 'Unit' and column_names[3] == 'Buying Rate (Deno 50 or less)' and column_names[4] == 'Buying Rate (Deno 50 or above)' and
                column_names[5] == 'Selling Rate' and column_names[6] == 'Prenium Rate'):

                    Unit_values = df['Unit']
                    Buying_less__values = df['Buying Rate (Deno 50 or less)']
                    Buying_above__values = df['Buying Rate (Deno 50 or above)']
                    Selling_Rate_values = df['Selling Rate']
                    Prenium_Rate_values = df['Prenium Rate']

                    non_float_columns = []

                    # Check for non-float values in each column
                    columns = {
                        'Buying Rate (Deno 50 or less)': Buying_less__values,
                        'Buying Rate (Deno 50 or above)': Buying_above__values,
                        'Selling Rate': Selling_Rate_values,
                        'Prenium Rate': Prenium_Rate_values
                    }
                    if not Unit_values.apply(lambda x: isinstance(x, int)).all():
                        non_float_columns.append('Unit')

                    for col_name, col_data in columns.items():
                        if not col_data.apply(lambda x: isinstance(x, float)).all():
                            non_float_columns.append(col_name)

                    if len(non_float_columns) == 0:
                        # All columns have float values
                        rate_master = FCYRateMaster.objects.create(
                            date=system_date,
                            time=system_time,
                            user=request.user.email
                        )

                        master_id = rate_master.pk

                        for index, row in df.iterrows():
                            FCYExchangeRate.objects.create(
                                masterid=master_id,
                                currency_code=row['Currency Code'],
                                currency_unit=row['Unit'],
                                currency=row['Currency'],
                                buying_rate_deno_50_or_less=row['Buying Rate (Deno 50 or less)'],
                                buying_rate_deno_50_or_above=row['Buying Rate (Deno 50 or above)'],
                                selling_rate=row['Selling Rate'],
                                premium_rate=row['Prenium Rate'],
                            )

                        messages.success(self.request, "Operation Successful!")
                        return redirect('/dashboard/')
                    else:
                        # Generate dynamic error message based on non-float columns
                        error_message = f"Invalid Excel File: The following column(s) contain non-float values - {', '.join(non_float_columns)}."
                        messages.error(self.request, error_message)
                        return render(request, self.template_name, {'form': form})
                else:
                    messages.error(
                        self.request, "Invalid Excel File, Please check the column header and try again!")
                    return render(request, self.template_name, {'form': form})
            else:
                return render(request, self.template_name, {'form': form})
        else:
            messages.error(self.request, "Unauthorized User")
            return render(request, self.template_name, {'form': form})


@csrf_exempt
def convert_to_words(request):
    if request.method == 'POST':
        total_equivalent_npr = request.POST.get('totalEquivalentNPR')

        # Calculate the amount in words
        total_equivalent_npr_to_words = (
            num2words(total_equivalent_npr, lang='en_IN') + ' only/-').title()

        data = {
            'totalEquivalentNPRToWords': total_equivalent_npr_to_words,  # Adjust the key name here
        }
        return JsonResponse(data)
    else:
        return JsonResponse({'totalEquivalentNPRToWords': ''})


@csrf_exempt
def getdenowiserate(request):
    if request.method == 'POST':
        try:
            deno = int(request.POST.get('denoValue'))
            currency = request.POST.get('selectedCurrency')

            try:
                last_object = FCYRateMaster.objects.latest('id')
                filtered_data = FCYExchangeRate.objects.filter(
                    masterid=last_object.pk, currency_code=currency).first()

                if filtered_data:
                    if deno <= 49:
                        denorate = filtered_data.buying_rate_deno_50_or_above
                    else:
                        denorate = filtered_data.premium_rate

                    data = {
                        'rate': denorate,
                    }
                    return JsonResponse(data)
                else:
                    return JsonResponse({'error': 'No exchange rate found for the given currency'}, status=404)
            except ObjectDoesNotExist:
                return JsonResponse({'error': 'No FCYRateMaster object found'}, status=404)

        except (ValueError, ObjectDoesNotExist) as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
        return JsonResponse({'rate': ''})


class FCYRequestView(LoginRequiredMixin, View):
    template_name = 'core/dashboard/myfcyrequest.html'

    def get(self, request):
        if(request.user.role == 0):
            preferred_branch_subquery = Branches.objects.filter(
                BranchCode=OuterRef('preferredBranch')).values('BranchName')[:1]
            customer_fullname = UserAccount.objects.annotate(
                full_name=ExpressionWrapper(
                    Concat(F('first_name'), Value(' '), F('last_name')),
                    output_field=CharField()
                )
            ).filter(email=OuterRef('enteredBy')).values('full_name')[:1]
            query_result = FCYExchangeRequestMaster.objects.annotate(
                prefBranchName=Subquery(preferred_branch_subquery),
                customer_fullname=Subquery(customer_fullname)
            ).exclude(status='Deleted').order_by('-refrenceid').values()
        elif(request.user.role == 1):
            preferred_branch_subquery = Branches.objects.filter(
                BranchCode=OuterRef('preferredBranch')).values('BranchName')[:1]
            customer_fullname = UserAccount.objects.annotate(
                full_name=ExpressionWrapper(
                    Concat(F('first_name'), Value(' '), F('last_name')),
                    output_field=CharField()
                )
            ).filter(email=OuterRef('enteredBy')).values('full_name')[:1]
            query_result = FCYExchangeRequestMaster.objects.annotate(
                prefBranchName=Subquery(preferred_branch_subquery),
                customer_fullname=Subquery(customer_fullname)
            ).exclude(status='Deleted').order_by('-refrenceid').filter(enteredBy = request.user.email).values()
        elif(request.user.role == 2):
            preferred_branch_subquery = Branches.objects.filter(
                BranchCode=OuterRef('preferredBranch')).values('BranchName')[:1]
            customer_fullname = UserAccount.objects.annotate(
                full_name=ExpressionWrapper(
                    Concat(F('first_name'), Value(' '), F('last_name')),
                    output_field=CharField()
                )
            ).filter(email=OuterRef('enteredBy')).values('full_name')[:1]
            query_result = FCYExchangeRequestMaster.objects.annotate(
                prefBranchName=Subquery(preferred_branch_subquery),
                customer_fullname=Subquery(customer_fullname)
            ).exclude(status='Deleted').order_by('-refrenceid').filter(preferredBranch = request.user.branch).values()
            
        else:
            messages.error(self.request, "Unauthorized User")
            return redirect('/dashboard/')

        return render(request, self.template_name, {'fcyrequest': query_result})

    def post(self, request, *args, **kwargs):
        return HttpResponse('POST request!')


class FCYRequestDetailView(LoginRequiredMixin, View):
    template_name = 'core/dashboard/detailfcyrequest.html'

    def get(self, request, id):
        preferred_branch_subquery = Branches.objects.filter(
            BranchCode=OuterRef('preferredBranch')).values('BranchName')[:1]
        customer_fullname = UserAccount.objects.annotate(
            full_name=ExpressionWrapper(
                Concat(F('first_name'), Value(' '), F('last_name')),
                output_field=CharField()
            )
        ).filter(email=OuterRef('enteredBy')).values('full_name')[:1]
        fcyrequest = FCYExchangeRequestMaster.objects.filter(id=id).annotate(
            prefBranchName=Subquery(preferred_branch_subquery),
            customer_fullname=Subquery(customer_fullname)
        ).exclude(status='Deleted').values().get()
        if request.user.email == fcyrequest['enteredBy'] or request.user.role == 0 or request.user.role == 2:
            fcydenodetails = FCYDenoMasterTable.objects.filter(
                masterid=id).order_by('currency', 'deno')

            # Create a dictionary to organize data by currency
            data_by_currency = defaultdict(list)
            # Group data by currency
            for item in fcydenodetails:
                currency = item.currency
                data_by_currency[currency].append({
                    'deno': item.deno,
                    'unit': item.unit,
                    'rate': item.rate,
                    'equivalent_npr': item.equivalentNPR
                })
            return render(request, self.template_name, {'fcyrequest': fcyrequest, 'data_by_currency': dict(data_by_currency)})
        else:
            messages.error(self.request, "Unauthorized User")
            return redirect('/dashboard/')
        

    def post(self, request, *args, **kwargs):
        return HttpResponse('POST request!')


class FCYRequestEditView(LoginRequiredMixin, View):
    template_name = 'core/dashboard/editfcyrequest.html'

    def get(self, request, id):
        preferred_branch_subquery = Branches.objects.filter(
            BranchCode=OuterRef('preferredBranch')).values('BranchName')[:1]
        customer_fullname = UserAccount.objects.annotate(
            full_name=ExpressionWrapper(
                Concat(F('first_name'), Value(' '), F('last_name')),
                output_field=CharField()
            )
        ).filter(email=OuterRef('enteredBy')).values('full_name')[:1]
        fcyrequest = FCYExchangeRequestMaster.objects.filter(id=id).annotate(
            prefBranchName=Subquery(preferred_branch_subquery),
            customer_fullname=Subquery(customer_fullname)
        ).exclude(status='Deleted').values().get()
        if request.user.email == fcyrequest['enteredBy'] or request.user.role == 0 or request.user.role == 2:
            fcydenodetails = FCYDenoMasterTable.objects.filter(masterid=id, status='Requested')
            form = DenoApprovedForm()
            return render(request, self.template_name, {'fcyrequest': fcyrequest, 'fcydenodetails': fcydenodetails, 'form':form})
        else:
            messages.error(self.request, "Unauthorized User")
            return redirect('/dashboard/')
        

    def post(self, request, *args, **kwargs):
        form = DenoApprovedForm(request.POST)
        id = kwargs.get('id')
        fcymaster = get_object_or_404(FCYExchangeRequestMaster, id=id)

        if fcymaster.enteredBy == request.user.email or request.user.role == 2:
            if form.is_valid():
                action = form.cleaned_data['action']
                if action == 'APPROVED':
                    fcymaster.status = 'Approved'
                    fcymaster.remarks = form.cleaned_data['remarks']
                    fcymaster.depositedby = form.cleaned_data['depositedby']
                    fcymaster.updatedBy = request.user.email if request.user.is_authenticated else ''
                    fcymaster.updateDate = current_datetime
                    fcymaster.save()
                    send_email_with_attachment(id)
                elif action == 'REJECTED':
                    fcymaster.status = 'Rejected'
                    fcymaster.remarks = form.cleaned_data['remarks']
                    fcymaster.depositedby = form.cleaned_data['depositedby']
                    fcymaster.updatedBy = request.user.email if request.user.is_authenticated else ''
                    fcymaster.updateDate = current_datetime
                    fcymaster.save()
                    send_email_rejection(id)
                else:
                    messages.error(request, "Operation Unsuccessful!")
                    return redirect(f'/edit/fcyexchangerequest/{id}') 


                messages.success(request, "Operation Successful!")
                return redirect('/fcyexchangerequest/') 
            else:
                messages.error(request, "Form data is invalid!")
                return redirect(f'/edit/fcyexchangerequest/{id}') 
        else:
            messages.error(request, "Unauthorized User!")
            return redirect('/fcyexchangerequest/')

@require_POST
def update_fcy_data(request, fcy_id, masterId):
    fcy = get_object_or_404(FCYDenoMasterTable, id=fcy_id)
    fcymaster = get_object_or_404(FCYExchangeRequestMaster, id=masterId)
    if 'deno' in request.POST:
        fcy.deno = request.POST['deno']
    if 'rate' in request.POST:
        fcy.rate = request.POST['rate']
    if 'unit' in request.POST:
        fcy.unit = request.POST['unit']
    if 'equivalentNPR' in request.POST:
        fcy.equivalentNPR = request.POST['equivalentNPR']
    if 'totalEquivalentNPR' in request.POST:
        fcymaster.totalEquivalentNPR = request.POST['totalEquivalentNPR']
    if 'totalEquivalentNPRToWords' in request.POST:
        fcymaster.totalEquivalentNPRToWords = request.POST['totalEquivalentNPRToWords']
    fcy.updatedBy = request.user.email if request.user.is_authenticated else ''
    fcy.updateDate = current_datetime
    fcymaster.updatedBy = request.user.email if request.user.is_authenticated else ''
    fcymaster.updateDate = current_datetime
    fcymaster.save()
    fcy.save()

    return JsonResponse({'message': 'Data updated successfully'})


def generate_pdf_receipt(request, id):
    exchnagedata = get_object_or_404(FCYExchangeRequestMaster, id=id)
    userdata = UserAccount.objects.get(email=exchnagedata.enteredBy)
    fcydenodetails = FCYDenoMasterTable.objects.filter(
        masterid=exchnagedata.id).order_by('currency', 'deno')

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    logo_path = find('images/logo.png')
    if logo_path:
        img_width = 130
        img_height = 45
        page_width, page_height = A4
        x_position = (page_width - img_width) / 2
        y_position = 785
        p.drawImage(logo_path, x_position, y_position,
                    width=img_width, height=img_height)

    p.setFillColor(colors.black)
    p.line(0, 770, 600, 770)

    text = "FOREIGN EXCHANGE ENCASHMENT RECEIPT"
    p.setFont("Times-Bold", 16)
    text_width = p.stringWidth(text)
    page_width, _ = A4
    x_position = (page_width - text_width) / 2
    p.setFillColor('#0366ae')
    p.drawString(x_position, 750, text)

    p.setFillColor(colors.black)
    p.setFont("Times-Roman", 12)
    p.drawString(50, 725, f"Receipt No: {exchnagedata.refrenceid}")
    p.drawString(50, 710, f"Date: {system_date}")

    left_margin = 50
    right_margin = A4[0] - 50

    text_content = (
        f'We hereby certify that we have purchased today foreign currency as mentioned below from M/S {userdata.first_name} {userdata.last_name}.'
    )

    text_x = left_margin
    text_y = 685
    text_width = right_margin - left_margin

    lines = []
    line = ''
    for word in text_content.split():
        if p.stringWidth(line + ' ' + word, "Times-Roman", 12) < text_width:
            line += f'{word} '
        else:
            lines.append(line)
            line = f'{word} '

    lines.append(line)

    p.setFont("Times-Roman", 12)
    p.setFillColor(colors.black)
    for line in lines:
        p.drawString(text_x, text_y, line.strip())
        text_y -= 15

    text_y -= 15
    table_height = (len(fcydenodetails)+2) * 15
    table_y = text_y - table_height
    table_data = [
        ['Currency', 'Deno', 'Unit', 'Rate', 'Equivalent NPR']
    ]

    for fcy in fcydenodetails:
        row = [
            fcy.currency[:3],
            fcy.deno,
            fcy.unit,
            fcy.rate,
            fcy.equivalentNPR
        ]
        table_data.append(row)

    table = Table(table_data, colWidths=100)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#d2a12a'),
        ('TEXTCOLOR', (0, 0), (-1, 0), '#0366ae'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Times-Roman'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige)
    ]))

    table.wrapOn(p, A4[0], A4[1])
    table.drawOn(p, 50, table_y)

    text_below_table_y = table_y - 15
    p.setFillColor(colors.black)
    p.setFont("Times-Bold", 12)
    p.drawString(370, text_below_table_y,
                 f"Total Equivalent NPR: {exchnagedata.totalEquivalentNPR}")
    text_below_table_y = table_y - 15

    # Draw the text below the table
    p.setFillColor(colors.black)
    p.setFont("Times-Roman", 12)
    text_width = A4[0] - 2 * 50
    text_object = p.beginText(50, text_below_table_y - 20)
    text_object.setFont("Times-Roman", 12)
    text_object.setTextOrigin(50, text_below_table_y - 20)
    text_object.setTextOrigin(50, text_below_table_y - 20)
    text_object.setTextOrigin(50, text_below_table_y - 20)
    text_object.textLines(
        f"NPR Amount in Words: {exchnagedata.totalEquivalentNPRToWords}")
    p.drawText(text_object)
    p.drawString(50, text_below_table_y - 60,
                 f"Customer's Signature:_______________________________")
    p.drawString(50, text_below_table_y - 160,
                 f"Signature & Stamp of the Bank")
    
    
    p.showPage()
    p.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f'{userdata.first_name}-{userdata.last_name}-{exchnagedata.date} FCY_Exchange_report.pdf')


def send_email_with_attachment(id):
    exchnagedata = get_object_or_404(FCYExchangeRequestMaster, id=id)
    userdata = UserAccount.objects.get(email=exchnagedata.enteredBy)

    # Generate the PDF using the function
    pdf_buffer = generate_pdf(userdata, exchnagedata)

    # Create an EmailMultiAlternatives object to include both HTML and text content
    email = EmailMultiAlternatives(
        subject='Foreign Exchange Encashment Receipt',
        from_email=settings.EMAIL_HOST_USER,  # Replace with your sender email address
        to=['shekhar.dhakal@jbbl.com.np', exchnagedata.enteredBy],  # Replace with recipient email address
    )

    # Attach the generated PDF file to the email
    email.attach(f'{userdata.first_name}-{userdata.last_name}-{exchnagedata.date} FCY_Exchange_report.pdf', pdf_buffer.getvalue(), 'application/pdf')

    # Email body (HTML content)
    html_content = render_to_string('core/email/encashment_receipt_template.html', {
        'username': f'{userdata.first_name} {userdata.last_name}',  
    })

    # Attach HTML content as an alternative
    email.attach_alternative(html_content, "text/html")

    # Send the email
    email.send()
    
def send_email_rejection(id):
    exchnagedata = get_object_or_404(FCYExchangeRequestMaster, id=id)
    userdata = UserAccount.objects.get(email=exchnagedata.enteredBy)
    email = EmailMultiAlternatives(
        subject='Foreign Currency Exchange Request Rejection',
        from_email=settings.EMAIL_HOST_USER,  
        to=['shekhar.dhakal@jbbl.com.np', exchnagedata.enteredBy], 
    )
    html_content = render_to_string('core/email/rejection_email.html', {
        'username': f'{userdata.first_name} {userdata.last_name}', 
        'remarks': exchnagedata.remarks,
    })
    email.attach_alternative(html_content, "text/html")
    email.send()
    
def generate_pdf(userdata, exchnagedata):
    fcydenodetails = FCYDenoMasterTable.objects.filter(
        masterid=exchnagedata.id).order_by('currency', 'deno')

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    logo_path = find('images/logo.png')
    if logo_path:
        img_width = 130
        img_height = 45
        page_width, page_height = A4
        x_position = (page_width - img_width) / 2
        y_position = 785
        p.drawImage(logo_path, x_position, y_position,
                    width=img_width, height=img_height)

    p.setFillColor(colors.black)
    p.line(0, 770, 600, 770)

    text = "FOREIGN EXCHANGE ENCASHMENT RECEIPT"
    p.setFont("Times-Bold", 16)
    text_width = p.stringWidth(text)
    page_width, _ = A4
    x_position = (page_width - text_width) / 2
    p.setFillColor('#0366ae')
    p.drawString(x_position, 750, text)

    p.setFillColor(colors.black)
    p.setFont("Times-Roman", 12)
    p.drawString(50, 730, f"Receipt No: {exchnagedata.refrenceid}")
    p.drawString(50, 715, f"Date: {system_date}")

    left_margin = 50
    right_margin = A4[0] - 50

    text_content = (
        f'We hereby certify that we have purchased today foreign currency as mentioned below from M/S {userdata.first_name} {userdata.last_name}.'
    )

    text_x = left_margin
    text_y = 685
    text_width = right_margin - left_margin

    lines = []
    line = ''
    for word in text_content.split():
        if p.stringWidth(line + ' ' + word, "Times-Roman", 12) < text_width:
            line += f'{word} '
        else:
            lines.append(line)
            line = f'{word} '

    lines.append(line)

    p.setFont("Times-Roman", 12)
    p.setFillColor(colors.black)
    for line in lines:
        p.drawString(text_x, text_y, line.strip())
        text_y -= 15

    text_y -= 15
    table_height = (len(fcydenodetails)+2) * 15
    table_y = text_y - table_height
    table_data = [
        ['Currency', 'Deno', 'Unit', 'Rate', 'Equivalent NPR']
    ]

    for fcy in fcydenodetails:
        row = [
            fcy.currency[:3],
            fcy.deno,
            fcy.unit,
            fcy.rate,
            fcy.equivalentNPR
        ]
        table_data.append(row)

    table = Table(table_data, colWidths=100)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#d2a12a'),
        ('TEXTCOLOR', (0, 0), (-1, 0), '#0366ae'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Times-Roman'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige)
    ]))

    table.wrapOn(p, A4[0], A4[1])
    table.drawOn(p, 50, table_y)

    text_below_table_y = table_y - 15
    p.setFillColor(colors.black)
    p.setFont("Times-Bold", 12)
    p.drawString(370, text_below_table_y,
                 f"Total Equivalent NPR: {exchnagedata.totalEquivalentNPR}")
    text_below_table_y = table_y - 15

    # Draw the text below the table
    p.setFillColor(colors.black)
    p.setFont("Times-Roman", 12)
    text_width = A4[0] - 2 * 50
    text_object = p.beginText(50, text_below_table_y - 20)
    text_object.setFont("Times-Roman", 12)
    text_object.setTextOrigin(50, text_below_table_y - 20)
    text_object.setTextOrigin(50, text_below_table_y - 20)
    text_object.setTextOrigin(50, text_below_table_y - 20)
    text_object.textLines(
        f"NPR Amount in Words: {exchnagedata.totalEquivalentNPRToWords}")
    p.drawText(text_object)
    p.drawString(50, text_below_table_y - 60,
                 f"Customer's Signature:_______________________________")
    p.drawString(50, text_below_table_y - 160,
                 f"Signature & Stamp of the Bank")
    p.drawString(50, 100, f"Note: This is electronically generated report. Please print this report and sign the document for further references.")

    p.showPage()
    p.save()
    buffer.seek(0)
    return buffer





from django.db import connection
from openpyxl.styles import Font

@login_required
def generate_xls_batch(request, id):
    exchnagedata = get_object_or_404(FCYExchangeRequestMaster, id=id)
    userdata = UserAccount.objects.get(email=exchnagedata.enteredBy)
    masterid = id
    with connection.cursor() as cursor:
        try:
            cursor.execute('CALL createbatchdatas(%s)', [masterid])
            results = cursor.fetchall()
        except Exception as e:
            pass

    raw_query = """
        SELECT 
            'FC1' AS BATCHNO,
            SHORTCURRENCY,
            currency,
            -1 * SUM(equivalentNPR) AS equivalentNPR,
            "MainCode",
            "CyCode",
            BRANCHCODE,
            '017' AS TRANCODE,
            -1 * FCYAMOUNT,
            Deno,
            Unit,
            Rate
        FROM (
            SELECT 
                LEFT(d.currency, 3) AS SHORTCURRENCY,
                d.currency,
                SUM(d."equivalentNPR") AS equivalentNPR,
                c."MainCode",
                c."CyCode",
                m."preferredBranch" AS BRANCHCODE,
                d."deno" as Deno,
                d."unit" as Unit,
                d."rate" as Rate,
                (d."deno" * d."unit") AS FCYAMOUNT,
                ROW_NUMBER() OVER (PARTITION BY d.currency) AS rn
            FROM public.core_fcydenomastertable AS d
            JOIN public.branches_branchtelloraccount AS c ON c."CyDesc" = LEFT(d.currency, 3)
            JOIN public.core_fcyexchangerequestmaster AS m ON m."id" = d.masterid
            WHERE d.masterid = %s
            GROUP BY d.currency, SHORTCURRENCY, c."MainCode", c."CyCode", m."preferredBranch", d."deno", d."unit", d."rate"
        ) AS subquery
        GROUP BY currency, SHORTCURRENCY, "MainCode", "CyCode", BRANCHCODE, FCYAMOUNT, Deno, Unit, Rate;
    """

    with connection.cursor() as cursor:
        cursor.execute(raw_query, [masterid])
        rows = cursor.fetchall()
    wb = Workbook()
    ws = wb.active  # Get the active sheet

    # Add some data to the Excel file
    ws['A1'] = 'BATCHNO'
    ws['B1'] = 'BRANCHCODE'
    ws['C1'] = 'MAINCODE'
    ws['D1'] = 'DESC1'
    ws['E1'] = 'DESC2'
    ws['F1'] = 'DESC3'
    ws['G1'] = 'AMOUNT'
    ws['H1'] = 'LCYAMOUNT'
    ws['I1'] = 'TRANCODE'
    
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
    data = [
        ('FC1',exchnagedata.preferredBranch,userdata.client_code,f'{userdata.company[:20]}', f'{exchnagedata.depositedby[:20]}',f'{exchnagedata.refrenceid}', exchnagedata.totalEquivalentNPR, exchnagedata.totalEquivalentNPR, '517'),
    ]

    for row in rows:
        row_data = (
            row[0],
            row[6],
            row[4],
            f'{userdata.company[:20]}',
            row[1], 
            f'{row[9]}' '*' f'{row[10]}' '*' f'{row[11]}',
            row[8], 
            row[3],
            row[7],   
           
        )
        data.append(row_data)
        
    for row_num, row_data in enumerate(data, start=2):
        ws.cell(row=row_num, column=1).value = row_data[0]
        ws.cell(row=row_num, column=2).value = row_data[1]
        ws.cell(row=row_num, column=3).value = row_data[2]
        ws.cell(row=row_num, column=4).value = row_data[3]
        ws.cell(row=row_num, column=5).value = row_data[4]
        ws.cell(row=row_num, column=6).value = row_data[5]
        ws.cell(row=row_num, column=7).value = row_data[6]
        ws.cell(row=row_num, column=8).value = row_data[7]
        ws.cell(row=row_num, column=9).value = row_data[8]

    # Set the response content type for an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Batch File-{system_date}.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response


class FCYExchnageRateView(LoginRequiredMixin, View):
    template_name = 'core/dashboard/exchangerate.html'

    def get(self, request):


        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        return HttpResponse('POST request!')
